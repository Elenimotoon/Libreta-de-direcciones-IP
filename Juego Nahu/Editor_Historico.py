from tkinter import *
#import tkinter as tk
from tkinter import ttk, messagebox
from PIL import ImageTk, Image
# Llamo las dependencia de Tkinter y Pillow que voy a utilizar
import json
# Esto nos ayuda a leer y manipular archivos JSON
from ast import literal_eval
# Esto es para solucionar el problema de los strings con valores
from time import sleep
# Y el tiempo por si quieren animar algunas cosas
from pathlib import Path
import os
# Para solucionar problemas (como el mío) que el sistema no lo detecte 😅
from openpyxl import Workbook

# TODO Ya no funciona para lo mismo
# Actualizo las columnas con los valores almacenados
def update(prueba):
    for i in prueba:
        trv.insert('', 'end', values=i)

# Si la checkbox está marcada la desmarca y si está desmarcada la marca
def toggleCheck(event):
    row_id = tv.identify_row(event.y)
    if not row_id == '' and not tv.item(row_id, "tags") == '':
        tag = tv.item(row_id, "tags")[0]
        tags = list(tv.item(row_id, "tags"))
        tags.remove(tag)
        tv.item(row_id, tags=tags)
        if tag == "checked":
            tv.item(row_id, tags="unchecked")
        elif tag == "unchecked":
            tv.item(row_id, tags="checked")

# Tomo los valores de las columnas y las paso a las casillas
# TODO A futuro editar esto
def tomar_columnas(event):
    row_id = trv.identify_row(event.y)
    if not row_id == '':
        item = trv.item(trv.focus())
        if tipo_de_caja.widgetName == "entry":
            caja_valores.set(item['values'][0])
        else:
            tipo_de_caja.delete("1.0", END)
            tipo_de_caja.insert(END, item['values'][0])

# Tomo los valores de las carpetas y las paso a las columnas
# Y pongo el tipo de caja para ese valor
def tomar_carpeta(event):
    row_id = tv.identify_row(event.y)
    datos_importantes = any("checked" == tags or "unchecked" == tags for tags in list(tv.item(row_id, "tags")))
    if not row_id == '' and not len(tv.item(row_id, "values")) == 0 and not datos_importantes:
        clave = tv.item(row_id, "text")
        valor = tv.item(row_id, "values")
        trv.delete(*trv.get_children())
        try:
            valor = literal_eval(valor[0].replace('\\', ''))
            if not isinstance(valor, (dict, list, set, tuple)):
                trv.insert('', 'end', text=clave+":", values=valor)
                actualizar_valores(clave, valor)
        except:
            trv.insert('', 'end', text=clave+":", values=valor)
            actualizar_valores(clave, valor)

# Añado los valores de las cajas a la columna
# TODO Cambiar columnas por carpetas
"""
def aniadir_nodo():
    titulo = caja_titulo.get()
    vida_max = caja_vida_max.get()
    historia = caja_historia.get()
    
    if titulo == '':
        titulo = None
    if not vida_max.isnumeric() or vida_max == "0":
        return messagebox.showerror("Error de lectura", "Por favor, revise los datos ingresados y vuelva a intentarlo")
    if historia == '':
        historia = None
    
    # Elimino los valores de las columnas actuales
    trv.delete(*trv.get_children())
    
    # prueba.append(['Titulo', titulo])
    # prueba.append(['Vida Máxima', int(vida_max)])
    # prueba.append(['Historia', historia])
    # Actualizo las columnas con los valores actuales
    # update(prueba)
    
    # Limpio las cajas
    caja_titulo.set('')
    caja_vida_max.set('')
    caja_historia.set('')
    messagebox.showinfo("Guardado", "¡Nodo guardado exitosamente!")
"""

# Actualizo los valores tanto para el JSON como para las columnas
# FIXME Aun sigo trabajando en esto
def actualizar_nodo():
    item = trv.item(trv.focus())
    
    if item['text'] == "":
        return messagebox.showerror("Error", "No se a proporcionado un nodo a actualizar")
    
    if tipo_de_caja.widgetName == "entry":
        valores = caja_valores.get()
        if not valores.isnumeric() or valores == "0":
            return messagebox.showerror("Error de lectura", "Por favor, ingrese un número mayor a 0 y vuelva a intentarlo")
        if messagebox.askyesno("Advertencia", f"¡Está apunto de cambiar los valores de {item['text'][:-1]}!\n\n¿Está seguro de hacerlo?"):
            trv.item(trv.focus(), values=valores)
            caja_valores.set('')
    else:
        valores = tipo_de_caja.get("1.0", 'end-1c')
        if valores == '': 'None'
        if messagebox.askyesno("Advertencia", f"¡Está apunto de cambiar los valores de {item['text'][:-1]}!\n\n¿Está seguro de hacerlo?"):
            trv.item(trv.focus(), values=valores)
            tipo_de_caja.delete("1.0", END)
"""
# Borra el nodo seleccionado
def borrar_nodo():
    item_id = trv.focus()
    item = trv.item(item_id)
    if len(item['values']) == 0:
        return messagebox.showerror("Error", "No se a proporcionado un nodo a eliminar")
    if messagebox.askyesno("Advertencia", "¡Si elimina este nodo no habrá vuelta atrás!\n\n¿Está seguro?"):
        tags = list(trv.get_children())
        tags.remove(item_id)
        # prueba.remove(item['values'])
        trv.delete(*trv.get_children())
        # update(prueba)
        print("Nodo eliminado...")
    return True
"""
# Pequeña prueba para volver los botones y entradas más dinámicas
def actualizar_valores(dinamico, entrada):
    etiqueta_dinamica.config(text=dinamico+':')
    
    # Entry(wrapper2, textvariable=caja_valores) if isinstance(prueba[0][1], int) else Text(wrapper2, height=5, width=20, font=("Arial", 10))
    
    # TODO Encontrar una mejor solución que solo usar palabra clave global
    global tipo_de_caja
    
    vacio = len(tipo_de_caja.grid_info())
    if vacio != 0:
        if tipo_de_caja.widgetName == "entry" and isinstance(entrada, int):
            tipo_de_caja = Entry(wrapper2, textvariable=caja_valores)
        elif tipo_de_caja.widgetName == "text" and isinstance(entrada, str):
            Text(wrapper2, height=5, width=20, font=("Arial", 10))
        else:
            tipo_de_caja.grid_remove()
            tipo_de_caja = Entry(wrapper2, textvariable=caja_valores) if isinstance(entrada, int) else Text(wrapper2, height=5, width=20, font=("Arial", 10))
    else:
        tipo_de_caja.grid_remove()
        tipo_de_caja = Entry(wrapper2, textvariable=caja_valores) if isinstance(entrada, int) else Text(wrapper2, height=5, width=20, font=("Arial", 10))
        
        # boton_aniadir = Button(wrapper2, text="Añadir Nodo", command=aniadir_nodo)
        boton_actualizar = Button(wrapper2, text="Actualizar Valor", command=actualizar_nodo)
        boton_guardar = Button(wrapper2, text="Guardar Cambios", command=lambda: print('Supongamos que se guardaron'))
        # boton_borrar = Button(wrapper2, text="Borrar Nodo", command=borrar_nodo)
        
        # boton_aniadir.grid(row=1, column=0, padx=5, pady=3)
        boton_actualizar.grid(row=1, column=0, padx=5, pady=3)
        boton_guardar.grid(row=1, column=2, padx=5, pady=3)
        # boton_borrar.grid(row=1, column=2, padx=5, pady=3)
    
    tipo_de_caja.grid(row=0, column=1, padx=5, pady=3)

# Me aseguro que el usuario quiera salir
def salir():
    advertencia = messagebox.askyesnocancel("Advertencia", "¡Está apunto de salir sin haber guardado previamente los cambios!\n\n¿De sea guardar los cambios antes de salir?")
    if advertencia:
        ventana.config(cursor="wait")
        ventana.update()
        sleep(3)
        ventana.config(cursor="")
        ventana.quit()
    elif advertencia is None:
        pass
    else:
        ventana.quit()

# TODO Agreguen el resto de ventanas pls, acá tienen la base
# Muchos de los datos necesarios están en ReglasEstructura.md
def ventana_reglas():
    reglas = Toplevel(ventana)
    reglas.geometry("250x250")
    reglas.title("Reglas")
    reglas.resizable(False, False)
    # reglas.group(ventana)   # reglas.wm_group(ventana)
    descripcion = Label(reglas, text="Estas son las reglas")
    descripcion.pack()
    cerrar = Button(reglas, text="Cerrar", command=lambda: reglas.destroy())
    cerrar.pack()

# Creo una ventana de Tkinter
ventana = Tk()
# Preparo las cajas para que puedan recibir valores
caja_valores = StringVar()

# Menu Bar
menubar = Menu(ventana)
archivo_menu = Menu(menubar, tearoff=0)
# Añado los botones que van a tener
archivo_menu.add_command(label="Formato Simple", command=lambda: print('Formato basico activado'))
archivo_menu.add_command(label="Formato Datallado", command=lambda: print('Formato pro activado'))
archivo_menu.add_command(label="Guardar", command=lambda: print('Guardando Archivo'))
# TODO Crear un botón para cerrar todas las ventanas que no sean la principal (Quizá con un grup)
# (Opcional) Añadir un botón de Save and exit
archivo_menu.add_separator()
archivo_menu.add_command(label="Guardar y Salir", command=ventana.quit)
archivo_menu.add_command(label="Salir", command=salir)
# Elijo el nombre del contenedor con los botones
menubar.add_cascade(label="Archivo", menu=archivo_menu)

info_menu = Menu(menubar, tearoff=0)
# Añado los botones que van a tener
info_menu.add_command(label="Reglas", command=ventana_reglas)
info_menu.add_command(label="Ejemplos", command=lambda: print("Mostando ejemplos"))
info_menu.add_command(label="Funcionamiento", command=lambda: print("Mostando funcionamiento"))
# Elijo el nombre del contenedor con los botones
menubar.add_cascade(label="Info", menu=info_menu)

# Label
# (El cuadro de la izquierda)
label_principal = Label(ventana)
label_principal.grid(row=0, column=0, padx=5, pady=3)

# (El cuadro de la derecha)
label_de_ayuda = Label(ventana)
label_de_ayuda.grid(row=0, column=1, padx=5, pady=3)

# Dentro
wrapper0 = LabelFrame(label_principal, text='Carpetitas')
wrapper1 = LabelFrame(label_de_ayuda, text='Visualizador')
wrapper2 = LabelFrame(label_de_ayuda, text='Modificador')

wrapper0.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper1.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper2.pack(fill="both", expand="yes", padx=20, pady=10)

img_ch = Image.open(f'{Path("Juego Nahu").absolute()}\\checked.png')
img_un = Image.open(f'{Path("Juego Nahu").absolute()}\\uncheked.png')
# Llamo las imágenes y les ajusto el tamaño
img_checkeado = ImageTk.PhotoImage(img_ch.resize((12, 12)))
img_descheckeado = ImageTk.PhotoImage(img_un.resize((12, 12)))

# Carpetitas
tv = ttk.Treeview(wrapper0)

tv.tag_configure('checked', image=img_checkeado)
# Establezco que cada carpetita con la etiqueta checked se le va a agregar la imagen marcada
# Y cada carpetita con la etiqueta unchecked se le va a agregarla imagen desmarcada
tv.tag_configure('unchecked', image=img_descheckeado)

# Prueba manual de carpetas anidadas con valores
main = tv.insert("", END, text="MAIN", values=[{}], open=1)
tv.insert(main, END, text="TITLE", values='g')
tv.insert(main, END, text="MAX_HEALTH", values=2)
health_options = tv.insert(main, END, text="HEALTH_OPTIONS", values=[{}])

tv.insert(health_options, END, text="NUMERIC_HEALTH", tags="unchecked", values=[True])
tv.insert(health_options, END, text="PERCENT_HEALTH", tags="unchecked", values=[False])
health_bar = tv.insert(health_options, END, text="HEALTH_BAR", values=[{}])

tv.insert(health_bar, END, text="HEALTH_BAR_DISPLAY", tags="unchecked", values=[True])
tv.insert(health_bar, END, text="LEFT_HEALTH_BAR_SYMBOL", values='g')
tv.insert(health_bar, END, text="RIGHT_HEALTH_BAR_SYMBOL", values='g')
tv.insert(health_bar, END, text="SIZE_SYMBOL", values=7)

tv.insert(health_options, END, text="HEALTH_FORMAT", values='g')
states = tv.insert(health_options, END, text="HIDDEN_STATS", tags="unchecked", values=[False])

menu = tv.insert("", END, text="MENU", values=[{}], open=1)
tv.insert(menu, END, text="PLAY", values='play')
tv.insert(menu, END, text="OPTIONS", values=[{}])
tv.insert(menu, END, text="CREDITS", values='prego')
tv.insert(menu, END, text="EXIT", values='exit')

history = tv.insert("", END, text="HISTORY", values=[{}], open=1)
tv.insert(history, END, text="PROLOG", values='ha')
tv.insert(history, END, text="CHAPTERS", values=[{}, {}])
tv.pack(side=LEFT)
print(tv.item(history))

# Prueba para cuando tenga que convertir datos de las carpetitas a columnas dinámicas
ejemplo = tv.item(states)['values'][0]
if not ejemplo == '':
    if ejemplo == 'True':
        tv.item(states, tags='checked')
    elif ejemplo == 'False':
        tv.item(states, tags='unchecked')
    elif isinstance(ejemplo, int):
        print('Número')
    elif ejemplo == 'None':
        print('No sé, bro')
print(tv.item(states))

# Visualizador
# Creo una caja con una columna que almacene los valores y 5 filas (Temporal)
trv = ttk.Treeview(wrapper1, columns=("estructura"), height="5")

trv.pack(side=LEFT)
#trv.place(x=0, y=0)
# FIXME Problema con el scroolbar horizontal:
# Al activarlo se arregla pero el height se rompe y viceversa
trv.heading("#0", text="Values")
# Nombre de la columna para el nombramiento de los valores
# Y nombre de la columna para los valores
trv.heading("estructura", text="Estructura")
trv.column("#0", width=50, minwidth=100)
# Modifico el width para poder hacer scrool horizaontal en las columnas
# (Posiblemente temporal)
trv.column("estructura", width=150, minwidth=200)

# Con doble click en las columnas llevo la info a las cajas
trv.bind('<Double 1>', tomar_columnas)
# Con doble click en las carpetas llevo la info a las columnas
tv.bind('<Double 1>', tomar_carpeta)
# Con un click en las carpetas con checkboxes se invierten los estados
tv.bind('<Button 1>', toggleCheck)
# Scrollbal Vertical
visualizador_scrollbar_y = ttk.Scrollbar(wrapper1, orient="vertical", command=trv.yview)
visualizador_scrollbar_y.pack(side=RIGHT, fill="y")
carpetitas_scrollbar_y = ttk.Scrollbar(wrapper0, orient="vertical", command=tv.yview)
carpetitas_scrollbar_y.pack(side=RIGHT, fill="y")

# Scrollbal Horizontal
visualizador_scrollbar_x = ttk.Scrollbar(wrapper1, orient="horizontal", command=trv.xview)
visualizador_scrollbar_x.pack(side=BOTTOM, fill="x")
carpetitas_scrollbar_x = ttk.Scrollbar(wrapper0, orient="horizontal", command=tv.xview)
carpetitas_scrollbar_x.pack(side=BOTTOM, fill="x")

trv.configure(yscrollcommand=visualizador_scrollbar_y.set, xscrollcommand=visualizador_scrollbar_x.set)
tv.configure(yscrollcommand=carpetitas_scrollbar_y.set, xscrollcommand=carpetitas_scrollbar_x.set)

# Actualizo las columnas para que carguen los valores que le añado por defecto
# update(prueba)

# Modificador
# Cajas temporales para ingresar texto
etiqueta_dinamica = Label(wrapper2)
etiqueta_dinamica.grid(row=0, column=0, padx=5, pady=3)

tipo_de_caja = Entry(wrapper2, textvariable=caja_valores)
"""
def json_a_exel():
    with open(f'{Path("Juego Nahu").absolute()}\\EscenariosPrincipales.json', 'r', encoding='utf8') as archivo:
        datos = json.load(archivo)
        pirncipal = list(datos.keys())
        print(list(datos[pirncipal[0]].items())[0])
        
        archivo = Workbook()
        hoja = archivo.active
        hoja.title = pirncipal[0]
        hoja2 = archivo.create_sheet(pirncipal[1])
        hoja3 = archivo.create_sheet(pirncipal[2])
        
        for clasificacion in [hoja, hoja2, hoja3]:
            d = list(datos[clasificacion.title].items())
            for items in d:
                if not isinstance(items[1], (int, str)):
                    clasificacion.append([items[0], str(items[1])])
                else:
                    clasificacion.append(items)
        
        archivo.save('datos.xlsx')
json_a_exel()
"""
ventana.title("Editor Histórico")
ventana.geometry('850x350')
ventana.resizable(False, False)
ventana.config(menu=menubar)

#ventana.protocol("WM_DELETE_WINDOW", salir)

ventana.mainloop()

##########
"""
# Prueba del sistema automático de jerarquía de carpetas (En progreso)
# PD: Luego si quieren les explico, de todos modos aun no está terminado, por eso se ve horrible

# import json
# with open(f'{Path("Juego Nahu").absolute()}\\EscenariosPrincipales.json', 'r', encoding='utf8') as archivo:
#     datos = json.load(archivo)
#     print(list(datos[list(datos.keys())[0]].items())[0][0])

si = {'MAIN': {'TITLE': 'Return: Mi travesía de volver a casa', 'MAX_HEALTH': 100, 'HEALTH_OPTIONS': {'NUMERIC_HEALTH': True, 'PERCENT_HEALTH': False, 'HEALTH_BAR': {'HEALTH_BAR_DISPLAY': True, 'LEFT_HEALTH_BAR_SYMBOL': '#', 'RIGHT_HEALTH_BAR_SYMBOL': '-', 'SIZE_SYMBOL': 20}, 'HEALTH_FORMAT': '{HEALTH_BAR}  {NUMERIC_HEALTH}', 'HIDDEN_STATS': False}}, 'MENU': {'PLAY': 'Jugar', 'OPTIONS': {'NAME': 'Opciones', 'DETAILS': {'VOLUME': 'Volumen', 'TEXT_SIZE': 16}}, 'CREDITS': 'Le agradezco a Python por dejarme programarle', 'EXIT': 'Salir'}, 'HISTORY': {'PROLOG': 'Una noche cualquiera, nuestro personaje se va a dormir, y luego de dormir, el siente que se despierta, pero se despierta en una realidad totalmente diferente, al ir a su armario, se encuentra con prendas de un origen desconocido y medianamente antiguas, las cuales se termina por poner, luego de salir de su hogar, se encuentra en un mundo distinto, en una aldea remota, luego de salir de esa aldea, nuestro personaje se pierde en un bosque, y a partir de ahí, comienza su aventura para poder volver a su mundo real\n¿Logrará hacerlo?', 'CHAPTERS': [{'CHAPTER_NAME': 'El bosque de las almas perdidas', 'INITIAL_HEALTH': {'SET_HEALTH': True, 'NEW_HEALTH': 100}, 'STORYTELLER': None, 'INTERNAL_VOICE': 'Acabo de perderme en este bosque, lo único que escucho mas alla de mis propios pasos son\ngritos de almas, desesperadas por volver a la vida, parece que estan sufriendo una horrible\n pesadilla sin poder descansar en paz\n\nMuchas de estas palabras, estan en idiomas que desconozco, ni siquiera estan hablando en\nEspañol o Ingles, podría llegar a decir que estas estan en un idioma tan antiguo como este\nlugar', 'PROTAGONIST_DIALOGUE': None, 'OPTIONAL_TEXT': "(Encontraste una nota, 'El sucumbir de las almas desesperadas')", 'PROTAGONIST_STATS': {'DAMAGE': None, 'HEALING': None, 'POISON': None, 'BLEED_OUT': None}}]}}

prueba = {'MAIN': 'Hola', 'MENU': '¿Qué tal?','HISTORY': {'Bien': 'si', "coso": [{"cosito": 'de cosoy'}, {"cosito2": 'de cosoy', 'codicia': 'si'}]}}

ventana = Tk()
tv = ttk.Treeview(ventana)

tv.insert("", END, "MAIN", text="MAIN", open=1)
tv.insert("", END, "MENU", text="MENU", open=1)
tv.insert("", END, "HISTORY", text="HISTORY", open=1)

def enrutador_principal(lista, buscar):
    def enrutador(estilo, valor):
        if estilo == valor:
            return [estilo]
        elif isinstance(estilo, dict):
            for k, v in estilo.items():
                p = enrutador(v, valor)
                if p:
                    return [k] + p
        elif isinstance(estilo, list):
            lst = estilo
            for i in range(len(lst)):
                p = enrutador(lst[i], valor)
                if p:
                    return [str(i)] + p
    try:
        return '/'.join(enrutador(lista, buscar)[:-1])
    except:
        return enrutador(lista, buscar)

ruta = []
ruta2 = []
candado = [False]
indexacion = []
cadena = []

def enrutador2():
    if not candado[0]:
        ruta2.extend(ruta)
    else:
        pass
    #print(ruta2)
    if not len(indexacion) == 0:
        for i in indexacion:
            if ruta[-1] == i[2]:
                mirar = cadena[0].split('/')
                print('/'.join(mirar))
                if len(mirar) > 2 and not mirar[-2] == i[1]:
                    mirar[-2:] = i[1:]
                    print('/'.join(mirar))
                # ruta.pop()
                # ruta.extend(i)
                # print(ruta[-3:])
                # indexacion.clear()
    else:
        print(cadena[0])
    print(ruta)
    # ruta2.clear()
    # ruta2.extend(ruta)
    # print(ruta2)

def crear_arbolito(arbolito):
    for i in list(arbolito.items()):
        if isinstance(i[1], (dict)):
            for q in i[1].items():
                tv.insert(i[0], END, q[0], text=q[0])
            ruta.append(i[0])
            crear_arbolito(i[1])
        
        elif isinstance(i[1], (list)):
            candado.clear()
            candado.append(True)
            ruta2.append(len(i[1])-1)
            for idx in range(len(i[1])):
                tv.insert(i[0], END, f'[{idx}]', text=f'[{idx}]')
                for exacto in list(i[1][idx].items()):
                    indexacion.append([i[0], str(idx), exacto[0]])
                    tv.insert(f'[{idx}]', END, exacto[0], text=exacto[0])
            for dentro in i[1]:
                crear_arbolito(dentro)
        else:
            ruta.append(i[0])
            if not candado[0]:
                ruta2.clear()
            cadena.append(enrutador_principal(prueba, i[1]))
            enrutador2()
            ruta.pop()
            cadena.pop()
            # enrutador2()
            print('\n')
            # ruta2.append(i[0])
            try:
                if isinstance(i[1], bool):
                    tv.insert(i[0], END, text=str(i[1]))
                else:
                    tv.insert(i[0], END, text=i[1])
            except:
                tv.insert(i[0], END, text=str(i[1]))

crear_arbolito(prueba)

tv.pack()

ventana.title("Editor Histórico")
ventana.geometry('750x350')

ventana.mainloop()
"""